import csv
import io
import re
from pathlib import Path

from as400_api import normalizar_debe_haber, AS400ApiError
from cuenta_tipos import resolver_codigo_cuenta


class CSVAsientoError(Exception):
    pass


ALIAS_COLUMNAS = {
    "cuenta": (
        "cuenta", "codigo", "codigocuenta", "codigocuenta", "codigo cuenta",
        "ncuenta", "n cuenta", "mccta", "adcta", "aicta", "acta", "ctacontable",
        "cuentacontable", "cuentacont", "numcuenta",
    ),
    "fecha": (
        "fecha", "date", "fec", "fechacontable", "fechaoperacion", "fechavalor",
        "foperacion", "fvalor",
    ),
    "importe": (
        "importe", "amount", "cantidad", "valor", "total", "adimp", "imp",
        "importelinea", "importeasiento",
    ),
    "debe_haber": (
        "debehaber", "debe haber", "debe_haber", "dh", "dh", "tipo",
        "movimiento", "signo", "addoh", "doh", "debeohaber", "ladocuenta",
    ),
    "debe": ("debe", "cargo", "importedebe", "debeimporte"),
    "haber": ("haber", "abono", "importehaber", "haberimporte"),
    "concepto": (
        "concepto", "descripcion", "desc", "texto", "glosa", "detalle",
        "comentario", "observaciones", "adcct", "mcnbr", "nombre", "titular",
        "programa", "conceptolinea",
    ),
}

EXTENSIONES_EXCEL = {".xlsx", ".xlsm", ".xls"}
EXTENSIONES_TEXTO_COLUMNAR = {
    ".csv", ".tsv", ".txt", ".dat", ".tab", ".psv",
}
EXTENSIONES_NORMA43 = {".n43", ".c43", ".csb", ".43"}


def _normalizar_encabezado(texto):
    return re.sub(r"[^a-z0-9]", "", str(texto or "").strip().lower())


def _mapear_columnas(encabezados):
    indice = {}
    normalizados = [_normalizar_encabezado(item) for item in encabezados]

    for campo, alias in ALIAS_COLUMNAS.items():
        for posicion, encabezado in enumerate(normalizados):
            if encabezado in alias:
                indice[campo] = posicion
                break

    for campo, alias in ALIAS_COLUMNAS.items():
        if campo in indice:
            continue

        alias_largos = sorted(
            (item for item in alias if len(item) >= 4),
            key=len,
            reverse=True,
        )

        for posicion, encabezado in enumerate(normalizados):
            if not encabezado:
                continue

            if any(
                clave == encabezado
                or clave in encabezado
                or encabezado in clave
                for clave in alias_largos
            ):
                indice[campo] = posicion
                break

    posicion_importe = indice.get("importe")

    if posicion_importe is not None:
        encabezado_importe = normalizados[posicion_importe]

        for campo, palabras in (
            ("debe", ("debe", "cargo")),
            ("haber", ("haber", "abono")),
        ):
            if indice.get(campo) == posicion_importe and not any(
                palabra in encabezado_importe
                for palabra in palabras
            ):
                indice.pop(campo, None)

    return indice


def decodificar_csv(contenido):
    if isinstance(contenido, str):
        return contenido

    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return contenido.decode(encoding)
        except UnicodeDecodeError:
            continue

    raise CSVAsientoError("No se pudo leer el archivo CSV (codificación no reconocida)")


def _detectar_delimitador(texto):
    muestra = "\n".join(texto.splitlines()[:5])

    try:
        dialecto = csv.Sniffer().sniff(muestra, delimiters=";,\t|")
        return dialecto.delimiter
    except csv.Error:
        primera_linea = texto.splitlines()[0] if texto else ""

        if primera_linea.count("|") >= max(
            primera_linea.count(";"),
            primera_linea.count(","),
            primera_linea.count("\t"),
        ):
            return "|"

        if primera_linea.count("\t") >= max(
            primera_linea.count(";"),
            primera_linea.count(","),
        ):
            return "\t"

        if primera_linea.count(";") >= primera_linea.count(","):
            return ";"

        return ","


def _leer_filas_csv(texto):
    texto = str(texto or "").strip()

    if not texto:
        raise CSVAsientoError("El archivo CSV está vacío")

    delimitador = _detectar_delimitador(texto)
    lector = csv.reader(io.StringIO(texto), delimiter=delimitador)
    filas = [
        [str(celda or "").strip() for celda in fila]
        for fila in lector
        if any(str(celda or "").strip() for celda in fila)
    ]

    if not filas:
        raise CSVAsientoError("El archivo CSV no contiene filas con datos")

    return filas, delimitador


def _parsear_importe(valor, permitir_cero=False):
    importe = _parsear_importe_bruto(valor)

    if importe is None:
        return None

    if importe == 0 and not permitir_cero:
        return None

    if importe < 0:
        return None

    return round(importe, 2)


def _parsear_importe_bruto(valor):
    texto = re.sub(r"[€$\s]", "", str(valor or "").strip())

    if not texto:
        return None

    negativo = texto.startswith("-") or texto.startswith("(") and texto.endswith(")")

    if texto.startswith("(") and texto.endswith(")"):
        texto = texto[1:-1]

    texto = texto.replace("-", "", 1) if texto.startswith("-") else texto

    if re.fullmatch(r"\d+", texto):
        importe = float(texto)
    elif "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
        importe = float(texto)
    elif "," in texto:
        importe = float(texto.replace(",", "."))
    else:
        try:
            importe = float(texto)
        except ValueError:
            raise CSVAsientoError(f"Importe no válido: {valor}") from None

    if negativo:
        importe = -abs(importe)

    return round(importe, 2)


def _parsear_fecha_csv(valor):
    texto = str(valor or "").strip()

    if not texto:
        return ""

    solo_digitos = re.sub(r"\D", "", texto)

    if len(solo_digitos) == 8:
        return f"{solo_digitos[:4]}-{solo_digitos[4:6]}-{solo_digitos[6:8]}"

    coincidencia = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", texto)

    if coincidencia:
        return texto

    coincidencia = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{4})", texto)

    if coincidencia:
        dia, mes, anio = coincidencia.groups()
        return f"{anio}-{mes.zfill(2)}-{dia.zfill(2)}"

    raise CSVAsientoError(f"Fecha no válida: {valor}")


def _normalizar_debe_haber_csv(valor):
    texto = str(valor or "").strip().upper()

    if not texto:
        raise CSVAsientoError("Falta indicar Debe o Haber")

    if texto in ("0", "D", "DEBE", "CARGO", "C"):
        return "D"

    if texto in ("1", "H", "HABER", "ABONO", "A"):
        return "H"

    try:
        return normalizar_debe_haber(texto)
    except AS400ApiError as error:
        raise CSVAsientoError(str(error)) from error


def _normalizar_cuenta_import(valor):
    texto = str(valor or "").strip()

    if not texto:
        return ""

    if re.fullmatch(r"\d+", texto):
        return texto.zfill(10)

    return texto


def _resolver_importe_y_lado(fila, columnas, numero_fila):
    if "debe" in columnas and "haber" in columnas:
        importe_debe = _parsear_importe(_valor_celda(fila, columnas["debe"]))
        importe_haber = _parsear_importe(_valor_celda(fila, columnas["haber"]))

        if importe_debe and importe_haber:
            raise CSVAsientoError(
                f"Fila {numero_fila}: no puede tener importe en Debe y Haber a la vez"
            )

        if importe_debe:
            return importe_debe, "D"

        if importe_haber:
            return importe_haber, "H"

    importe_bruto = _parsear_importe_bruto(_valor_celda(fila, columnas.get("importe")))

    if importe_bruto is None or importe_bruto == 0:
        raise CSVAsientoError(f"Fila {numero_fila}: falta el importe")

    debe_haber = None
    valor_dh = _valor_celda(fila, columnas.get("debe_haber"))

    if valor_dh != "":
        debe_haber = _normalizar_debe_haber_csv(valor_dh)

    if importe_bruto < 0:
        importe = abs(importe_bruto)

        if debe_haber is None:
            debe_haber = "H"
        else:
            debe_haber = "H" if debe_haber == "D" else "D"
    else:
        importe = importe_bruto

        if debe_haber is None:
            raise CSVAsientoError(f"Fila {numero_fila}: falta Debe/Haber")

    return importe, debe_haber


def _valor_celda(fila, indice):
    if indice is None or indice >= len(fila):
        return ""

    return fila[indice]


def _valores_fila_por_encabezado(encabezados, fila):
    valores = {}

    for posicion, encabezado in enumerate(encabezados):
        clave = str(encabezado or f"columna_{posicion + 1}").strip()
        valores[clave] = _valor_celda(fila, posicion)

    return valores


def _columnas_detectadas(encabezados, columnas):
    detectadas = []

    for campo in ("cuenta", "fecha", "importe", "debe_haber", "debe", "haber", "concepto"):
        posicion = columnas.get(campo)

        if posicion is None:
            continue

        detectadas.append({
            "campo": campo,
            "indice": posicion,
            "columna": encabezados[posicion] if posicion < len(encabezados) else "",
        })

    return detectadas


def _diagnostico_columnar(encabezados, columnas, datos):
    return {
        "columnas_detectadas": _columnas_detectadas(encabezados, columnas),
        "filas_total": len(datos),
        "filas_validas": 0,
        "filas_descartadas": 0,
        "errores": [],
        "cuentas_no_encontradas": [],
        "previsualizacion": [],
        "carga_parcial": False,
    }


def _registrar_previsualizacion(diagnostico, item, limite=12):
    if len(diagnostico["previsualizacion"]) < limite:
        diagnostico["previsualizacion"].append(item)


def _validar_cuenta_plan(linea, cuentas_plan, numero_fila, diagnostico):
    if cuentas_plan is None:
        return linea

    codigos_validos = {
        str(cuenta.get("codigo", "")).strip()
        for cuenta in cuentas_plan
        if str(cuenta.get("codigo", "")).strip()
    }
    cuenta_resuelta = resolver_codigo_cuenta(linea.get("cuenta", ""), cuentas_plan)

    if cuenta_resuelta and cuenta_resuelta in codigos_validos:
        linea = dict(linea)
        linea["cuenta"] = cuenta_resuelta
        return linea

    diagnostico["cuentas_no_encontradas"].append({
        "fila": numero_fila,
        "cuenta": linea.get("cuenta", ""),
        "concepto": linea.get("concepto", ""),
        "importe": linea.get("importe", 0),
        "debe_haber": linea.get("debe_haber", ""),
    })

    return None


def _es_csv_estructurado(columnas):
    return "cuenta" in columnas and (
        "importe" in columnas
        or ("debe" in columnas and "haber" in columnas)
    )


def _parsear_linea_estructurada(fila, columnas, numero_fila):
    cuenta = _normalizar_cuenta_import(_valor_celda(fila, columnas.get("cuenta")))

    if not cuenta:
        raise CSVAsientoError(f"Fila {numero_fila}: falta la cuenta contable")

    importe, debe_haber = _resolver_importe_y_lado(fila, columnas, numero_fila)

    fecha = ""

    if "fecha" in columnas:
        try:
            fecha = _parsear_fecha_csv(_valor_celda(fila, columnas["fecha"]))
        except CSVAsientoError:
            fecha = ""

    concepto = _valor_celda(fila, columnas.get("concepto")).strip()

    if not concepto:
        concepto = f"Línea importada ({numero_fila})"

    return {
        "cuenta": cuenta,
        "fecha": fecha,
        "importe": importe,
        "debe_haber": debe_haber,
        "concepto": concepto,
    }


def _limpiar_filas_tabla(filas):
    limpias = []

    for fila in filas:
        valores = [str(celda or "").strip() for celda in fila]

        if any(valores):
            limpias.append(valores)

    return limpias


def _leer_tabla_desde_excel(contenido):
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise CSVAsientoError(
            "Para importar Excel instala la dependencia openpyxl"
        ) from error

    libro = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    hoja = libro.active
    filas = _limpiar_filas_tabla(hoja.iter_rows(values_only=True))
    libro.close()

    if not filas:
        raise CSVAsientoError("El archivo Excel está vacío")

    return filas


def _es_fila_cabecera(fila):
    columnas = _mapear_columnas(fila)
    return _es_csv_estructurado(columnas)


def parsear_tabla_asiento(filas, delimitador=None, cuentas_plan=None):
    filas = _limpiar_filas_tabla(filas)

    if not filas:
        raise CSVAsientoError("El archivo no contiene filas con datos")

    indice_cabecera = 0

    for posicion, fila in enumerate(filas[:5]):
        if _es_fila_cabecera(fila):
            indice_cabecera = posicion
            break

    encabezados = filas[indice_cabecera]
    columnas = _mapear_columnas(encabezados)
    datos = filas[indice_cabecera + 1:]
    diagnostico = _diagnostico_columnar(encabezados, columnas, datos)

    if _es_csv_estructurado(columnas):
        lineas = []

        for indice, fila in enumerate(datos, start=indice_cabecera + 2):
            valores = _valores_fila_por_encabezado(encabezados, fila)

            try:
                linea = _parsear_linea_estructurada(fila, columnas, indice)
                linea = _validar_cuenta_plan(linea, cuentas_plan, indice, diagnostico)
            except CSVAsientoError as error:
                diagnostico["errores"].append({
                    "fila": indice,
                    "mensaje": str(error),
                    "valores": valores,
                })
                _registrar_previsualizacion(diagnostico, {
                    "fila": indice,
                    "estado": "error",
                    "mensaje": str(error),
                    "valores": valores,
                })
                continue

            if not linea:
                _registrar_previsualizacion(diagnostico, {
                    "fila": indice,
                    "estado": "cuenta_no_encontrada",
                    "mensaje": "Cuenta contable no encontrada en el plan",
                    "valores": valores,
                })
                continue

            if linea["importe"] > 0:
                lineas.append(linea)
                _registrar_previsualizacion(diagnostico, {
                    "fila": indice,
                    "estado": "ok",
                    "linea": linea,
                    "valores": valores,
                })

        diagnostico["filas_validas"] = len(lineas)
        diagnostico["filas_descartadas"] = (
            len(diagnostico["errores"])
            + len(diagnostico["cuentas_no_encontradas"])
        )
        diagnostico["carga_parcial"] = (
            diagnostico["filas_validas"] > 0
            and diagnostico["filas_descartadas"] > 0
        )

        if not lineas:
            return {
                "modo": "datos",
                "motivo": "sin_lineas_validas",
                "lineas": [],
                "num_filas": len(datos),
                "encabezados": encabezados,
                "delimitador": delimitador,
                "resumen": resumen_csv_para_ia(encabezados, datos),
                "diagnostico": diagnostico,
            }

        return {
            "modo": "lineas",
            "lineas": lineas,
            "num_filas": len(lineas),
            "encabezados": encabezados,
            "delimitador": delimitador,
            "resumen": resumen_csv_para_ia(encabezados, datos),
            "diagnostico": diagnostico,
        }

    _registrar_previsualizacion(diagnostico, {
        "fila": indice_cabecera + 2,
        "estado": "datos",
        "mensaje": "No se detectaron las columnas mínimas para cargar líneas",
        "valores": _valores_fila_por_encabezado(encabezados, datos[0]) if datos else {},
    })

    return {
        "modo": "datos",
        "lineas": [],
        "num_filas": len(datos),
        "encabezados": encabezados,
        "delimitador": delimitador,
        "resumen": resumen_csv_para_ia(encabezados, datos),
        "diagnostico": diagnostico,
    }


def parsear_csv_asiento(contenido):
    texto = decodificar_csv(contenido)
    filas, delimitador = _leer_filas_csv(texto)
    resultado = parsear_tabla_asiento(filas, delimitador=delimitador)
    resultado["formato"] = "columnar"
    return resultado


def resumen_csv_para_ia(encabezados, filas, max_filas=40):
    lineas = [
        "Datos del CSV adjunto:",
        f"Columnas: {', '.join(encabezados)}",
        "",
    ]

    for indice, fila in enumerate(filas[:max_filas], start=1):
        valores = []

        for posicion, encabezado in enumerate(encabezados):
            valor = fila[posicion] if posicion < len(fila) else ""
            valores.append(f"{encabezado}={valor}")

        lineas.append(f"Fila {indice}: " + "; ".join(valores))

    if len(filas) > max_filas:
        lineas.append(f"... ({len(filas) - max_filas} filas más)")

    return "\n".join(lineas)


def _es_extension_excel(filename):
    extension = Path(str(filename or "")).suffix.lower()
    return extension in EXTENSIONES_EXCEL


def _es_extension_norma43(filename):
    extension = Path(str(filename or "")).suffix.lower()
    return extension in EXTENSIONES_NORMA43


def parsear_archivo_importacion(
    contenido,
    filename=None,
    cuenta_banco=None,
    cuentas_plan=None,
):
    from norma43_asiento import es_fichero_norma43, parsear_norma43_asiento

    if _es_extension_excel(filename):
        filas = _leer_tabla_desde_excel(contenido)
        resultado = parsear_tabla_asiento(filas, cuentas_plan=cuentas_plan)
        resultado.setdefault("formato", "excel")
        return resultado

    texto = decodificar_csv(contenido)
    filas = []
    delimitador = None

    if texto.strip():
        filas, delimitador = _leer_filas_csv(texto)

        if filas and _es_fila_cabecera(filas[0]):
            resultado = parsear_tabla_asiento(
                filas,
                delimitador=delimitador,
                cuentas_plan=cuentas_plan,
            )
            resultado.setdefault("formato", "columnar")
            return resultado

    if _es_extension_norma43(filename) or (texto.strip() and es_fichero_norma43(texto)):
        resultado = parsear_norma43_asiento(
            contenido,
            cuenta_banco=cuenta_banco,
            cuentas_plan=cuentas_plan,
        )
        resultado.setdefault("formato", "norma43")
        return resultado

    if filas:
        resultado = parsear_tabla_asiento(
            filas,
            delimitador=delimitador,
            cuentas_plan=cuentas_plan,
        )
        resultado.setdefault("formato", "columnar")
        return resultado

    raise CSVAsientoError(
        "No se reconoció el formato del archivo. "
        "Usa columnas cuenta/importe/debe_haber, Excel encolumnado o Norma 43."
    )
