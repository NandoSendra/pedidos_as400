import re
import time
from pathlib import Path
from threading import Lock

from config import Config
from cuenta_tipos import _normalizar_texto, normalizar_iva_porcentaje

_lock = Lock()
_cache = {
    "loaded_at": 0.0,
    "mtime": None,
    "pgc_por_codigo": {},
    "subcuentas_por_codigo": {},
    "subcuentas_por_padre": {},
}


def plan_contable_file_path():
    return Path(Config.APP_PLAN_CONTABLE_IA_FILE)


def _celda_texto(valor):
    if valor is None:
        return ""

    texto = str(valor).strip()

    if texto.lower() in {"none", "nan"}:
        return ""

    return texto


def _merge_textos(*valores, separador="; "):
    partes = []
    vistos = set()

    for valor in valores:
        for parte in re.split(r"[;|,]", str(valor or "")):
            texto = parte.strip()

            if not texto:
                continue

            clave = _normalizar_texto(texto)

            if clave in vistos:
                continue

            vistos.add(clave)
            partes.append(texto)

    return separador.join(partes)


def _codigo_significativo(codigo):
    texto = re.sub(r"\D", "", str(codigo or "").strip())

    if not texto:
        return ""

    return texto.rstrip("0") or texto


def _codigos_relacionados(codigo_a, codigo_b):
    a = _codigo_significativo(codigo_a)
    b = _codigo_significativo(codigo_b)

    if not a or not b:
        return False

    return a == b or a.startswith(b) or b.startswith(a)


def _buscar_pgc(codigo, indice_pgc):
    codigo = re.sub(r"\D", "", str(codigo or "").strip())

    if not codigo:
        return None

    mejor = None
    mejor_longitud = 0

    for codigo_pgc, registro in indice_pgc.items():
        if codigo.startswith(codigo_pgc) and len(codigo_pgc) > mejor_longitud:
            mejor = registro
            mejor_longitud = len(codigo_pgc)

    return mejor


def _buscar_subcuenta(codigo, nombre, indice_sub, indice_padre):
    codigo = re.sub(r"\D", "", str(codigo or "").strip())
    nombre_norm = _normalizar_texto(nombre)

    for codigo_sub, registro in indice_sub.items():
        if _codigos_relacionados(codigo, codigo_sub):
            return registro

    padre = codigo[:3] if len(codigo) >= 3 else codigo
    candidatos = indice_padre.get(padre, [])

    if not candidatos or not nombre_norm:
        return None

    mejor = None
    mejor_puntuacion = 0

    for registro in candidatos:
        puntuacion = 0
        claves = _normalizar_texto(registro.get("palabras_clave", ""))

        for palabra in re.split(r"[^a-z0-9]+", nombre_norm):
            if len(palabra) < 3:
                continue

            if palabra in claves:
                puntuacion += 12
            elif palabra in _normalizar_texto(registro.get("descripcion", "")):
                puntuacion += 8

        if puntuacion > mejor_puntuacion:
            mejor_puntuacion = puntuacion
            mejor = registro

    return mejor if mejor_puntuacion >= 8 else None


def _cargar_plan_contable():
    ruta = plan_contable_file_path()

    if not ruta.is_file():
        return {
            "pgc_por_codigo": {},
            "subcuentas_por_codigo": {},
            "subcuentas_por_padre": {},
        }

    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            "pgc_por_codigo": {},
            "subcuentas_por_codigo": {},
            "subcuentas_por_padre": {},
        }

    mtime = ruta.stat().st_mtime

    with _lock:
        if (
            _cache["pgc_por_codigo"]
            and _cache["mtime"] == mtime
            and time.time() - _cache["loaded_at"] < 3600
        ):
            return {
                "pgc_por_codigo": dict(_cache["pgc_por_codigo"]),
                "subcuentas_por_codigo": dict(_cache["subcuentas_por_codigo"]),
                "subcuentas_por_padre": {
                    clave: list(valor)
                    for clave, valor in _cache["subcuentas_por_padre"].items()
                },
            }

    workbook = load_workbook(ruta, read_only=True, data_only=True)
    pgc_por_codigo = {}
    subcuentas_por_codigo = {}
    subcuentas_por_padre = {}

    if "PGC oficial" in workbook.sheetnames:
        hoja = workbook["PGC oficial"]

        for fila in hoja.iter_rows(min_row=2, values_only=True):
            codigo = _celda_texto(fila[0] if fila else "")

            if not codigo or not codigo.isdigit():
                continue

            pgc_por_codigo[codigo] = {
                "codigo": codigo,
                "descripcion": _celda_texto(fila[1]),
                "grupo": _celda_texto(fila[2]),
                "nombre_grupo": _celda_texto(fila[3]),
                "naturaleza": _celda_texto(fila[5]),
                "palabras_clave": _celda_texto(fila[6]),
                "palabras_excluyentes": _celda_texto(fila[7]),
                "contrapartidas": _celda_texto(fila[8]),
                "reglas": _celda_texto(fila[9]),
                "origen": _celda_texto(fila[10]),
            }

    if "Subcuentas sugeridas" in workbook.sheetnames:
        hoja = workbook["Subcuentas sugeridas"]

        for fila in hoja.iter_rows(min_row=2, values_only=True):
            codigo = _celda_texto(fila[0] if fila else "")

            if not codigo or not codigo.isdigit():
                continue

            padre = _celda_texto(fila[2])
            registro = {
                "codigo": codigo,
                "descripcion": _celda_texto(fila[1]),
                "padre": padre,
                "palabras_clave": _celda_texto(fila[3]),
                "palabras_excluyentes": _celda_texto(fila[4]),
                "contrapartidas": _celda_texto(fila[5]),
                "iva_orientativo": _celda_texto(fila[6]),
                "centro_coste": _celda_texto(fila[7]),
                "automatizacion": _celda_texto(fila[8]),
                "notas": _celda_texto(fila[9]),
            }
            subcuentas_por_codigo[codigo] = registro

            if padre:
                subcuentas_por_padre.setdefault(padre, []).append(registro)

    workbook.close()

    with _lock:
        _cache["loaded_at"] = time.time()
        _cache["mtime"] = mtime
        _cache["pgc_por_codigo"] = pgc_por_codigo
        _cache["subcuentas_por_codigo"] = subcuentas_por_codigo
        _cache["subcuentas_por_padre"] = subcuentas_por_padre

    return {
        "pgc_por_codigo": pgc_por_codigo,
        "subcuentas_por_codigo": subcuentas_por_codigo,
        "subcuentas_por_padre": subcuentas_por_padre,
    }


def plan_contable_disponible():
    datos = _cargar_plan_contable()
    return bool(datos["pgc_por_codigo"] or datos["subcuentas_por_codigo"])


def get_plan_contable_info():
    ruta = plan_contable_file_path()
    datos = _cargar_plan_contable()

    return {
        "archivo": str(ruta),
        "existe": ruta.is_file(),
        "disponible": plan_contable_disponible(),
        "num_pgc": len(datos["pgc_por_codigo"]),
        "num_subcuentas": len(datos["subcuentas_por_codigo"]),
    }


def _aplicar_registro_plan(cuenta, pgc=None, subcuenta=None):
    cuenta = dict(cuenta)
    fuentes = []

    if pgc:
        fuentes.append("pgc")
        cuenta["pgc_codigo"] = pgc.get("codigo")
        cuenta["pgc_descripcion"] = pgc.get("descripcion")
        cuenta["pgc_grupo"] = pgc.get("grupo")
        cuenta["pgc_nombre_grupo"] = pgc.get("nombre_grupo")
        cuenta["pgc_naturaleza"] = pgc.get("naturaleza")
        cuenta["palabras_clave"] = _merge_textos(
            cuenta.get("palabras_clave"),
            pgc.get("palabras_clave"),
        )
        cuenta["palabras_excluyentes"] = _merge_textos(
            cuenta.get("palabras_excluyentes"),
            pgc.get("palabras_excluyentes"),
        )
        cuenta["contrapartidas_habituales"] = _merge_textos(
            cuenta.get("contrapartidas_habituales"),
            pgc.get("contrapartidas"),
        )
        cuenta["reglas_ia"] = _merge_textos(
            cuenta.get("reglas_ia"),
            pgc.get("reglas"),
        )

        if not cuenta.get("grupo_pgc") and pgc.get("grupo"):
            cuenta["grupo_pgc"] = pgc.get("grupo")

    if subcuenta:
        fuentes.append("subcuenta")
        cuenta["subcuenta_sugerida"] = subcuenta.get("codigo")
        cuenta["palabras_clave"] = _merge_textos(
            cuenta.get("palabras_clave"),
            subcuenta.get("palabras_clave"),
            subcuenta.get("descripcion"),
        )
        cuenta["palabras_excluyentes"] = _merge_textos(
            cuenta.get("palabras_excluyentes"),
            subcuenta.get("palabras_excluyentes"),
        )
        cuenta["contrapartidas_habituales"] = _merge_textos(
            cuenta.get("contrapartidas_habituales"),
            subcuenta.get("contrapartidas"),
        )
        cuenta["reglas_ia"] = _merge_textos(
            cuenta.get("reglas_ia"),
            subcuenta.get("notas"),
            subcuenta.get("automatizacion"),
        )

        if subcuenta.get("centro_coste"):
            cuenta["centro_coste_orientativo"] = subcuenta.get("centro_coste")

        iva_orientativo = normalizar_iva_porcentaje(subcuenta.get("iva_orientativo"))

        if iva_orientativo is not None and cuenta.get("iva_porcentaje") is None:
            cuenta["iva_porcentaje"] = iva_orientativo

    if fuentes:
        cuenta["plan_contable_fuente"] = "+".join(fuentes)

    return cuenta


def consolidar_cuenta_con_plan(cuenta, datos_plan=None):
    datos_plan = datos_plan or _cargar_plan_contable()

    if not datos_plan["pgc_por_codigo"] and not datos_plan["subcuentas_por_codigo"]:
        return dict(cuenta)

    codigo = str(cuenta.get("codigo", "")).strip()
    nombre = str(cuenta.get("nombre", "")).strip()
    pgc = _buscar_pgc(codigo, datos_plan["pgc_por_codigo"])
    subcuenta = _buscar_subcuenta(
        codigo,
        nombre,
        datos_plan["subcuentas_por_codigo"],
        datos_plan["subcuentas_por_padre"],
    )

    if not pgc and not subcuenta:
        return dict(cuenta)

    return _aplicar_registro_plan(cuenta, pgc=pgc, subcuenta=subcuenta)


def consolidar_cuentas_con_plan_ia(cuentas):
    datos_plan = _cargar_plan_contable()

    if not datos_plan["pgc_por_codigo"] and not datos_plan["subcuentas_por_codigo"]:
        return list(cuentas)

    return [
        consolidar_cuenta_con_plan(cuenta, datos_plan=datos_plan)
        for cuenta in cuentas
    ]
